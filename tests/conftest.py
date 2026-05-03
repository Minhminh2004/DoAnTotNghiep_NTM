import gc
import time
from pathlib import Path
from datetime import datetime

import pytest
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app import app as flask_app


TEST_DB_URL = (
    "mssql+pyodbc://DESKTOP-33J7KC7/DATN_NTM_TEST"
    "?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
)

REPORT_DIR = Path("reports")
test_results = []


@pytest.fixture
def app():
    flask_app.config["TESTING"] = True
    return flask_app


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def engine():
    e = create_engine(TEST_DB_URL, future=True)
    try:
        yield e
    finally:
        e.dispose()
        gc.collect()


def pytest_runtest_setup(item):
    item.start_time = time.time()
    item.start_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if report.when == "call":
        test_results.append({
            "test_name": item.name,
            "status": "PASS" if report.passed else "FAIL",
            "start_time": getattr(item, "start_datetime", ""),
            "duration_seconds": round(time.time() - getattr(item, "start_time", time.time()), 4),
        })


def pytest_sessionfinish(session, exitstatus):
    try:
        REPORT_DIR.mkdir(parents=True, exist_ok=True)

        excel_file = REPORT_DIR / f"test_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Report"

        ws.append(["Tên test", "Trạng thái", "Thời gian bắt đầu", "Thời gian chạy (giây)"])

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")

        for result in test_results:
            ws.append([
                result["test_name"],
                result["status"],
                result["start_time"],
                result["duration_seconds"],
            ])

        for row in range(2, ws.max_row + 1):
            status_cell = ws[f"B{row}"]
            if status_cell.value == "PASS":
                status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                status_cell.fill = PatternFill("solid", fgColor="FFC7CE")

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[letter].width = max_len + 3

        wb.save(excel_file)

        print(f"\n Đã tạo báo cáo test: {excel_file.resolve()}")

    except Exception as e:
        print(f"\n Không tạo được báo cáo Excel: {e}")